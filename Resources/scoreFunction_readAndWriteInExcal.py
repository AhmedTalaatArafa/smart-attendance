from difflib import SequenceMatcher
from openpyxl import load_workbook
import os

def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()

def write_in_Excel(path,list_of_name,no_tutorial):
    file_name=path
    workbook = load_workbook(filename=file_name)
    spreadsheet = workbook.active

    total_row = spreadsheet.max_row

    expected = []
    for curr_name in range(0,len(list_of_name)):
        flag =0
        for row_cursr in range(1,total_row+1):
            if spreadsheet.cell(row = row_cursr,column= 1).value == list_of_name[curr_name]:
                spreadsheet.cell(row = row_cursr,column= no_tutorial+1 ).value = 1 
                flag=1
                break
        if flag ==0:
            expected.append(list_of_name[curr_name])

    for curr_name in range(0,len(expected)):
        big =[]
        for row_cursr in range(2,total_row+1):
            big.append(similar(spreadsheet.cell(row = row_cursr,column= 1).value,expected[curr_name]))
        index_large_number=big.index(max(big))            
        if big[index_large_number] > 0.85:
            spreadsheet.cell(row = index_large_number+2,column= no_tutorial+1 ).value = 1 
    
    output_list = []
    for row_cursr in range(2,total_row+1):
        if  spreadsheet.cell(row = row_cursr,column= no_tutorial+1 ).value == 1:
            output_list.append(spreadsheet.cell(row = row_cursr,column= 1).value)

    os.remove(file_name)
    workbook.save(file_name)

names =["Ahmed Tarek","Ahmned saeid","Armed osama","notammed Hassan","Ahmed Eanzouri","Ahmed Talaat"]
write_in_Excel("F:/study/image processing/excel/test.xlsx",names,4)

