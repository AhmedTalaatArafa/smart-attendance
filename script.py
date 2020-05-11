import matplotlib.pyplot as plt
import matplotlib.patches as patches
import numpy as np
import random
import mxnet as mx
from mxnet import gluon
from utilities.iam_dataset import IAMDataset, resize_image, crop_image, crop_handwriting_page
from utilities.expand_bounding_box import expand_bounding_box
from utilities.word_and_line_segmentation import SSD as WordSegmentationNet, predict_bounding_boxes
from utilities.word_to_line import sort_bbs_line_by_line, crop_line_images
from utilities.handwriting_line_recognition import Network as HandwritingRecognitionNet, handwriting_recognition_transform
from utilities.handwriting_line_recognition import decode as decoder_handwriting, alphabet_encoding
from utilities.beam_search import ctcBeamSearch
import cv2
from difflib import SequenceMatcher
from openpyxl import load_workbook
import os

ctx = mx.cpu()

def get_arg_max(prob):
    '''
    The greedy algorithm convert the output of the handwriting recognition network
    into strings.
    '''
    arg_max = prob.topk(axis=2).asnumpy()
    return decoder_handwriting(arg_max)[0]

def get_beam_search(prob, width=5):
    possibilities = ctcBeamSearch(prob.softmax()[0].asnumpy(), alphabet_encoding, None, width)
    return possibilities[0]

output_list = []


def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()

def write_in_Excel(path,list_of_name,no_tutorial):
    file_name=path
    workbook = load_workbook(filename=file_name)
    spreadsheet = workbook.active

    total_row = spreadsheet.max_row

    expected = []
    for curr_name in range(0,len(list_of_name)):
        flag =0
        for row_cursr in range(1,total_row+1):
            if spreadsheet.cell(row = row_cursr,column= 1).value == list_of_name[curr_name]:
                spreadsheet.cell(row = row_cursr,column= no_tutorial+1 ).value = 1 
                flag=1
                break
        if flag ==0:
            expected.append(list_of_name[curr_name])

    for curr_name in range(0,len(expected)):
        big =[]
        for row_cursr in range(2,total_row+1):
            big.append(similar(spreadsheet.cell(row = row_cursr,column= 1).value,expected[curr_name]))
        index_large_number=big.index(max(big))            
        if big[index_large_number] > 0.5:
            spreadsheet.cell(row = index_large_number+2,column= no_tutorial+1 ).value = 1 

    
    for row_cursr in range(2,total_row+1):
        if  spreadsheet.cell(row = row_cursr,column= no_tutorial+1 ).value == 1:
            output_list.append(spreadsheet.cell(row = row_cursr,column= 1).value)
    

    os.remove(file_name)
    workbook.save(file_name)
#

#
def raw(text):
    escape_dict={'\a':r'\a',
        '\b':r'\b',
        '\c':r'\c',
        '\f':r'\f',
        '\n':r'\n',
        '\r':r'\r',
        '\t':r'\t',
        '\v':r'\v',
        '\'':r'\'',
        '\"':r'\"',
        '\0':r'\0',
        '\1':r'\1',
        '\2':r'\2',
        '\3':r'\3',
        '\4':r'\4',
        '\5':r'\5',
        '\6':r'\6',
        '\7':r'\7',
        '\8':r'\8',
        '\9':r'\9',
         '\256': r'\256'} # notice this line is the first 3 digits of the resolution

    for k in escape_dict:
        if text.find(k) > -1:
            text = text.replace(k, escape_dict[k])

    return text
#


File = open("Data.txt","r+")

Tutorial_no = int(File.readline())

Excel_Sheet_Path = File.readline()

Photo_File_Path = File.readline()


Excel_Sheet_Path = Excel_Sheet_Path.rstrip("\n")

#Excel_Sheet_Path = raw(Excel_Sheet_Path)


Photo_File_Path = Photo_File_Path.rstrip("\n")

#Photo_File_Path = raw(Photo_File_Path)

File.truncate(0)

#

#########################################

img = cv2.imread(Photo_File_Path,0)
ret,img = cv2.threshold(img,127,255,cv2.THRESH_BINARY)
top,bottom = (img.shape[0])//10 , (img.shape[0])//10
left,right = (img.shape[1])//10 , (img.shape[1])//10
color = [255, 255, 255]
image = cv2.copyMakeBorder( img, top, bottom, left, right, cv2.BORDER_CONSTANT,value=color)

word_segmentation_net = WordSegmentationNet(2, ctx=ctx)
word_segmentation_net.load_parameters("model_checkpoint\ssd_word.params")
word_segmentation_net.hybridize()


min_c = 0.1
overlap_thres = 0.1
topk = 600
predicted_bb = predict_bounding_boxes(word_segmentation_net, image, min_c, overlap_thres, topk, ctx)

line_bbs = sort_bbs_line_by_line(predicted_bb, y_overlap=0.4)
line_images = crop_line_images(image, line_bbs)




handwriting_line_recognition_net = HandwritingRecognitionNet(rnn_hidden_states=512,
                                                             rnn_layers=2, ctx=ctx, max_seq_len=160)
handwriting_line_recognition_net.load_parameters("model_checkpoint\handwriting_line_amazon.params", ctx=ctx)
handwriting_line_recognition_net.hybridize()

line_image_size = (60, 800)
form_character_prob = []
for i, line_image in enumerate(line_images):
    top,bottom = (line_image.shape[0])//4 , (line_image.shape[0])//4
    left,right = (line_image.shape[1])//4, (line_image.shape[1])//4
    color = [255, 255, 255]
    line_image = cv2.copyMakeBorder(line_image, top, bottom, left, right, cv2.BORDER_CONSTANT,value=color)
    line_image = handwriting_recognition_transform(line_image, line_image_size)
    line_character_prob = handwriting_line_recognition_net(line_image.as_in_context(ctx))
    form_character_prob.append(line_character_prob)
    
final_decoded=[]
final_text=[]

for j, line_character_probs in enumerate(form_character_prob):
        decoded_line_bs = get_beam_search(line_character_probs)
        final_text.append(decoded_line_bs)


for i in final_text:
    File.write(i+'\n')
#############################################

write_in_Excel(Excel_Sheet_Path, final_text, Tutorial_no)

for i in output_list:
    File.write(i+'\n')

output_list.clear()

File.close()
